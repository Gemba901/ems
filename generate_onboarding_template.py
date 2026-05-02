import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = openpyxl.Workbook()

# ── Styles ──────────────────────────────────────────────────────────────────
DARK_BLUE   = "1E3A5F"
MID_BLUE    = "2E6DA4"
LIGHT_BLUE  = "D6E4F0"
ACCENT      = "F0A500"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F5F5F5"
RED         = "C0392B"
GREEN       = "27AE60"
BORDER_GRAY = "CCCCCC"

def make_border(style="thin"):
    s = Side(style=style, color=BORDER_GRAY)
    return Border(left=s, right=s, top=s, bottom=s)

def header_font(size=11, bold=True, color=WHITE):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color="000000"):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

# ════════════════════════════════════════════════════════════════════════════
# Sheet 1 — Cover / Instructions
# ════════════════════════════════════════════════════════════════════════════
ws_cover = wb.active
ws_cover.title = "Instructions"
ws_cover.sheet_view.showGridLines = False
ws_cover.column_dimensions["A"].width = 4
ws_cover.column_dimensions["B"].width = 28
ws_cover.column_dimensions["C"].width = 55
ws_cover.column_dimensions["D"].width = 18
ws_cover.column_dimensions["E"].width = 4

# Banner
ws_cover.merge_cells("B2:D2")
ws_cover.merge_cells("B3:D3")
ws_cover.merge_cells("B4:D4")

banner_title = ws_cover["B2"]
banner_title.value = "GEMBA EMS"
banner_title.font = Font(name="Calibri", size=28, bold=True, color=WHITE)
banner_title.fill = PatternFill("solid", fgColor=DARK_BLUE)
banner_title.alignment = center()

banner_sub = ws_cover["B3"]
banner_sub.value = "Employee Onboarding Data Template"
banner_sub.font = Font(name="Calibri", size=16, bold=False, color=WHITE)
banner_sub.fill = PatternFill("solid", fgColor=DARK_BLUE)
banner_sub.alignment = center()

banner_date = ws_cover["B4"]
banner_date.value = "Please read instructions before filling in the Employee Data sheet"
banner_date.font = Font(name="Calibri", size=10, italic=True, color="D6E4F0")
banner_date.fill = PatternFill("solid", fgColor=DARK_BLUE)
banner_date.alignment = center()

ws_cover.row_dimensions[2].height = 40
ws_cover.row_dimensions[3].height = 30
ws_cover.row_dimensions[4].height = 22

# Instructions section
instructions = [
    ("HOW TO USE THIS TEMPLATE", None, True, DARK_BLUE, WHITE, 13),
    ("1. Navigate to the \"Employee Data\" sheet (tab at the bottom).", None, False, None, "000000", 10),
    ("2. Fill in one employee per row. Do NOT alter column headers.", None, False, None, "000000", 10),
    ("3. Required fields are marked with  ★  in the header row of the data sheet.", None, False, None, "000000", 10),
    ("4. Optional fields can be left blank.", None, False, None, "000000", 10),
    ("5. Use the exact values listed in the \"Reference Lists\" sheet for dropdown fields.", None, False, None, "000000", 10),
    ("6. Delete any example rows before submitting.", None, False, None, "000000", 10),
    ("7. Save the file and send it to your Gemba representative.", None, False, None, "000000", 10),
]

row = 6
for text, _, bold, bg, fg, size in instructions:
    ws_cover.merge_cells(f"B{row}:D{row}")
    cell = ws_cover[f"B{row}"]
    cell.value = text
    cell.font = Font(name="Calibri", size=size, bold=bold, color=fg)
    cell.alignment = left()
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
        ws_cover.row_dimensions[row].height = 26
    else:
        ws_cover.row_dimensions[row].height = 20
        cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY if (row % 2 == 0) else WHITE)
    row += 1

# Field legend
row += 1
ws_cover.merge_cells(f"B{row}:D{row}")
leg_head = ws_cover[f"B{row}"]
leg_head.value = "FIELD LEGEND"
leg_head.font = Font(name="Calibri", size=13, bold=True, color=WHITE)
leg_head.fill = PatternFill("solid", fgColor=DARK_BLUE)
leg_head.alignment = left()
ws_cover.row_dimensions[row].height = 26
row += 1

legends = [
    ("★  Required", "Must be provided for every employee. Rows missing required fields will be rejected."),
    ("(optional)", "May be left blank. Can be updated later inside the system."),
    ("Dropdown", "Must match one of the exact values in the Reference Lists sheet."),
]
for label, desc in legends:
    ws_cover[f"B{row}"].value = label
    ws_cover[f"B{row}"].font = Font(name="Calibri", size=10, bold=True, color=RED)
    ws_cover[f"B{row}"].alignment = left()
    ws_cover[f"C{row}"].value = desc
    ws_cover[f"C{row}"].font = body_font()
    ws_cover[f"C{row}"].alignment = left()
    ws_cover.row_dimensions[row].height = 18
    for col in ["B", "C", "D"]:
        ws_cover[f"{col}{row}"].fill = PatternFill("solid", fgColor=LIGHT_GRAY if row % 2 == 0 else WHITE)
    row += 1

# Contact
row += 1
ws_cover.merge_cells(f"B{row}:D{row}")
contact = ws_cover[f"B{row}"]
contact.value = "Questions? Contact your Gemba representative at info@gembapms.com"
contact.font = Font(name="Calibri", size=10, italic=True, color=MID_BLUE)
contact.alignment = center()
ws_cover.row_dimensions[row].height = 20

# ════════════════════════════════════════════════════════════════════════════
# Sheet 2 — Employee Data
# ════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Employee Data")
ws.sheet_view.showGridLines = False
ws.freeze_panes = "A3"  # freeze first 2 header rows

# Column definitions: (header_line1, header_line2, width, required, note)
columns = [
    # Personal Information
    ("#",               "",                     5,   False, "Auto — leave blank"),
    ("First Name",      "★ Required",           18,  True,  "Legal first name"),
    ("Last Name",       "★ Required",           18,  True,  "Legal last name"),
    ("Email Address",   "★ Required",           28,  True,  "Work or personal email — must be unique"),
    ("Phone Number",    "optional",             18,  False, "Format: +CountryCode Number"),
    ("Date of Birth",   "optional",             16,  False, "Format: YYYY-MM-DD"),
    ("National ID /\nPassport No.",  "optional", 20, False, "Government-issued ID number"),
    ("Address",         "optional",             32,  False, "Full residential address"),
    # Employment Details
    ("Department",      "optional  (Dropdown)", 22,  False, "See Reference Lists sheet"),
    ("Role / Job Title","★ Required (Dropdown)",22,  True,  "See Reference Lists sheet"),
    ("Employment Type", "optional  (Dropdown)", 20,  False, "Full-Time / Part-Time / Contract / Intern"),
    ("Employee Status", "optional  (Dropdown)", 20,  False, "Active / Onboarding / Suspended"),
    ("Supervisor Email","optional",             28,  False, "Email of direct supervisor (must already exist in system)"),
    # System
    ("Notes",           "optional",             35,  False, "Any additional info for the HR team"),
]

# Section headers spanning groups
SECTIONS = {
    1: ("PERSONAL INFORMATION", 8),   # cols 1-8 (A–H)
    9: ("EMPLOYMENT DETAILS",   5),   # cols 9-13 (I–M)
    14: ("NOTES", 1),
}

# Row 1 — section bands
for start_col, (label, span) in SECTIONS.items():
    end_col = start_col + span - 1
    start_ltr = get_column_letter(start_col)
    end_ltr   = get_column_letter(end_col)
    ws.merge_cells(f"{start_ltr}1:{end_ltr}1")
    cell = ws[f"{start_ltr}1"]
    cell.value = label
    cell.font  = Font(name="Calibri", size=11, bold=True, color=WHITE)
    cell.fill  = PatternFill("solid", fgColor=MID_BLUE if label != "PERSONAL INFORMATION" else DARK_BLUE)
    cell.alignment = center()

ws.row_dimensions[1].height = 22

# Row 2 — column headers
for i, (h1, h2, width, required, _) in enumerate(columns):
    col = i + 1
    ltr = get_column_letter(col)
    ws.column_dimensions[ltr].width = width

    cell = ws.cell(row=2, column=col)
    cell.value = f"{h1}\n{h2}" if h2 else h1
    cell.font  = Font(name="Calibri", size=10, bold=True,
                      color=WHITE if required else "D6E4F0")
    cell.fill  = PatternFill("solid", fgColor=DARK_BLUE)
    cell.alignment = center()
    cell.border = make_border()

ws.row_dimensions[2].height = 38

# Data rows 3-52 (50 employees)
dv_employment = DataValidation(
    type="list",
    formula1='"Full-Time,Part-Time,Contract,Intern"',
    allow_blank=True, showDropDown=False
)
dv_status = DataValidation(
    type="list",
    formula1='"Active,Onboarding,Suspended"',
    allow_blank=True, showDropDown=False
)
ws.add_data_validation(dv_employment)
ws.add_data_validation(dv_status)

for r in range(3, 53):
    bg = LIGHT_GRAY if r % 2 == 0 else WHITE
    for c in range(1, len(columns) + 1):
        cell = ws.cell(row=r, column=c)
        cell.fill   = PatternFill("solid", fgColor=bg)
        cell.font   = body_font()
        cell.border = make_border()
        cell.alignment = left()
    # Row number
    ws.cell(row=r, column=1).value = r - 2
    ws.cell(row=r, column=1).alignment = center()
    ws.cell(row=r, column=1).font = body_font(color="888888")

    # Dropdowns
    dv_employment.add(ws.cell(row=r, column=11))
    dv_status.add(ws.cell(row=r, column=12))

# Example row (row 3) — pre-filled sample
sample = [
    1, "Jane", "Doe", "jane.doe@company.com", "+254 700 000 000",
    "1990-05-15", "12345678", "123 Main St, Nairobi",
    "Engineering", "Software Engineer", "Full-Time", "Active",
    "manager@company.com", "Transfer from HQ"
]
for c, val in enumerate(sample, 1):
    cell = ws.cell(row=3, column=c)
    cell.value = val
    cell.font  = Font(name="Calibri", size=10, italic=True, color="555555")

# Row 3 label in col A
ws.cell(row=3, column=1).value = "EXAMPLE"
ws.cell(row=3, column=1).font  = Font(name="Calibri", size=8, bold=True, color=RED)

# Required column highlights
req_cols = [2, 3, 4, 10]  # firstName, lastName, email, role
for c in req_cols:
    cell = ws.cell(row=2, column=c)
    cell.fill = PatternFill("solid", fgColor="1A3050")

# ════════════════════════════════════════════════════════════════════════════
# Sheet 3 — Reference Lists
# ════════════════════════════════════════════════════════════════════════════
ws_ref = wb.create_sheet("Reference Lists")
ws_ref.sheet_view.showGridLines = False
ws_ref.column_dimensions["A"].width = 4
ws_ref.column_dimensions["B"].width = 28
ws_ref.column_dimensions["C"].width = 28
ws_ref.column_dimensions["D"].width = 28
ws_ref.column_dimensions["E"].width = 28
ws_ref.column_dimensions["F"].width = 4

ref_row_state = [2]

def ref_section(title, items, col_start="B"):
    ref_row = ref_row_state[0]
    col_end_map = {"B": "E"}
    col_end = col_end_map.get(col_start, "E")
    ws_ref.merge_cells(f"{col_start}{ref_row}:{col_end}{ref_row}")
    h = ws_ref[f"{col_start}{ref_row}"]
    h.value = title
    h.font  = Font(name="Calibri", size=11, bold=True, color=WHITE)
    h.fill  = PatternFill("solid", fgColor=DARK_BLUE)
    h.alignment = left()
    ws_ref.row_dimensions[ref_row].height = 22
    ref_row += 1
    for item in items:
        cell = ws_ref[f"{col_start}{ref_row}"]
        cell.value = item
        cell.font  = body_font()
        cell.alignment = left()
        cell.fill  = PatternFill("solid", fgColor=LIGHT_GRAY if ref_row % 2 == 0 else WHITE)
        cell.border = make_border()
        ws_ref.row_dimensions[ref_row].height = 18
        ref_row += 1
    ref_row += 1  # spacer
    ref_row_state[0] = ref_row

ref_section("EMPLOYMENT TYPE — Valid values for the 'Employment Type' column", [
    "Full-Time",
    "Part-Time",
    "Contract",
    "Intern",
])
ref_section("EMPLOYEE STATUS — Valid values for the 'Employee Status' column", [
    "Active",
    "Onboarding",
    "Suspended",
])
ref_section("DEPARTMENT — Common examples (update to match your organisation)", [
    "Engineering",
    "Human Resources",
    "Finance",
    "Operations",
    "Sales",
    "Marketing",
    "Customer Support",
    "Legal",
    "Executive",
    "(Add your departments below)",
])
ref_section("ROLE — Common examples (update to match your organisation)", [
    "Software Engineer",
    "HR Manager",
    "Accountant",
    "Operations Manager",
    "Sales Representative",
    "Marketing Specialist",
    "Support Agent",
    "Legal Counsel",
    "Director",
    "Chief Executive Officer",
    "(Add your roles below)",
])
ref_section("DATE FORMAT", [
    "Use YYYY-MM-DD   e.g. 1990-05-15",
])
ref_section("PHONE FORMAT", [
    "Include country code   e.g. +254 700 123 456",
])

# ════════════════════════════════════════════════════════════════════════════
# Tab colours
# ════════════════════════════════════════════════════════════════════════════
ws_cover.sheet_properties.tabColor = DARK_BLUE
ws.sheet_properties.tabColor       = MID_BLUE
ws_ref.sheet_properties.tabColor   = ACCENT

# ════════════════════════════════════════════════════════════════════════════
# Save
# ════════════════════════════════════════════════════════════════════════════
out = "/home/rodgers/Desktop/Gemba/ems-system/Gemba_EMS_Employee_Onboarding_Template.xlsx"
wb.save(out)
print(f"Saved: {out}")
